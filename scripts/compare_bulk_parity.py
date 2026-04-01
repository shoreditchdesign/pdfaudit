from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def _normalize_bucket(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def main(control_json_path: str, report_xlsx_path: str, output_csv_path: str) -> int:
    control_rows = json.loads(Path(control_json_path).read_text())
    workbook = load_workbook(Path(report_xlsx_path), data_only=True)
    worksheet = workbook["Sheet1"]
    headers = [cell.value for cell in worksheet[1]]
    header_index = {str(name): idx + 1 for idx, name in enumerate(headers) if name is not None}

    report_rows: dict[str, dict[str, object]] = {}
    for row_idx in range(2, worksheet.max_row + 1):
        original_url = worksheet.cell(row_idx, header_index["Original URL"]).value
        if not original_url:
            continue
        report_rows[str(original_url)] = {
            "pdf_name": worksheet.cell(row_idx, header_index["PDF Name"]).value,
            "axes_status": worksheet.cell(row_idx, header_index["Axes Audit Status"]).value,
            "pdfua": worksheet.cell(row_idx, header_index["PDF/ UA Reults"]).value,
            "wcag": worksheet.cell(row_idx, header_index["WCAG Reults"]).value,
            "pdfua_notes": worksheet.cell(row_idx, header_index["PDF/UA Notes"]).value,
            "wcag_notes": worksheet.cell(row_idx, header_index["WCAG Notes"]).value,
            "adobe_status": worksheet.cell(row_idx, header_index["Adobe Acrobat Audit Status"]).value,
            "final_url": worksheet.cell(row_idx, header_index["Final URL"]).value,
            "retrieval_category": worksheet.cell(row_idx, header_index["Retrieval Category"]).value,
            "overall_result": worksheet.cell(row_idx, header_index["Overall Result"]).value,
            "failure_summary": worksheet.cell(row_idx, header_index["Failure Summary"]).value,
        }

    output_path = Path(output_csv_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "Source Row Number",
                "PDF Name",
                "URL",
                "Control Axes Status",
                "Our Axes Status",
                "Axes Match",
                "Control PDF/UA",
                "Our PDF/UA",
                "PDF/UA Match",
                "Control WCAG",
                "Our WCAG",
                "WCAG Match",
                "Final URL",
                "Retrieval Category",
                "Overall Result",
                "Failure Summary",
                "Control PDF/UA Notes",
                "Our PDF/UA Notes",
                "Control WCAG Notes",
                "Our WCAG Notes",
                "Adobe Acrobat Audit Status",
            ],
        )
        writer.writeheader()

        for item in control_rows:
            report_row = report_rows.get(item["url"], {})
            control_axes = _normalize_bucket(item.get("control_axes_status"))
            our_axes = _normalize_bucket(report_row.get("axes_status"))
            control_pdfua = _normalize_bucket(item.get("control_pdfua"))
            our_pdfua = _normalize_bucket(report_row.get("pdfua"))
            control_wcag = _normalize_bucket(item.get("control_wcag"))
            our_wcag = _normalize_bucket(report_row.get("wcag"))
            writer.writerow(
                {
                    "Source Row Number": item.get("source_row_number"),
                    "PDF Name": item.get("pdf_name"),
                    "URL": item.get("url"),
                    "Control Axes Status": control_axes,
                    "Our Axes Status": our_axes,
                    "Axes Match": str(bool(control_axes and control_axes == our_axes)),
                    "Control PDF/UA": control_pdfua,
                    "Our PDF/UA": our_pdfua,
                    "PDF/UA Match": str(bool(control_pdfua and control_pdfua == our_pdfua)),
                    "Control WCAG": control_wcag,
                    "Our WCAG": our_wcag,
                    "WCAG Match": str(bool(control_wcag and control_wcag == our_wcag)),
                    "Final URL": report_row.get("final_url"),
                    "Retrieval Category": report_row.get("retrieval_category"),
                    "Overall Result": report_row.get("overall_result"),
                    "Failure Summary": report_row.get("failure_summary"),
                    "Control PDF/UA Notes": item.get("control_pdfua_notes"),
                    "Our PDF/UA Notes": report_row.get("pdfua_notes"),
                    "Control WCAG Notes": item.get("control_wcag_notes"),
                    "Our WCAG Notes": report_row.get("wcag_notes"),
                    "Adobe Acrobat Audit Status": report_row.get("adobe_status"),
                }
            )

    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1], sys.argv[2], sys.argv[3]))
