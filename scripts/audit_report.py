from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


STATUS_STYLE_MAP = {
    "PASS": {"fill": "DCFCE7", "font": "166534"},
    "Pass": {"fill": "DCFCE7", "font": "166534"},
    "Needs Review": {"fill": "FEF3C7", "font": "92400E"},
    "Pending Adobe": {"fill": "E5E7EB", "font": "4B5563"},
    "COMPLETE": {"fill": "DBEAFE", "font": "1D4ED8"},
    "Complete": {"fill": "DBEAFE", "font": "1D4ED8"},
    "FAIL": {"fill": "FEE2E2", "font": "B91C1C"},
    "Fail": {"fill": "FEE2E2", "font": "B91C1C"},
    "HARD_404": {"fill": "FEE2E2", "font": "B91C1C"},
    "SOFT_404": {"fill": "FEE2E2", "font": "B91C1C"},
    "REQUEST_ERROR": {"fill": "FEE2E2", "font": "B91C1C"},
    "BLOCKED": {"fill": "FDE68A", "font": "92400E"},
    "Blocked": {"fill": "FDE68A", "font": "92400E"},
    "ALT_LANDING_PAGE_NON_PDF": {"fill": "FDE68A", "font": "92400E"},
    "REVIEW_REQUIRED": {"fill": "FEF3C7", "font": "92400E"},
    "NEEDS_MANUAL_REVIEW": {"fill": "FEF3C7", "font": "92400E"},
    "API_UNAVAILABLE": {"fill": "E5E7EB", "font": "4B5563"},
    "N/A": {"fill": "F3F4F6", "font": "6B7280"},
    "Not Started": {"fill": "F3F4F6", "font": "6B7280"},
}

SHEET_SPECS = [
    (
        "Summary",
        [
            "PDF Name",
            "Original URL",
            "Final URL",
            "Date Audited",
            "Overall Result",
            "PDF/ UA Reults",
            "WCAG Reults",
            "Adobe Acrobat Audit Status",
            "PDF/UA Notes",
            "WCAG Notes",
            "Failure Summary",
        ],
        {"Overall Result", "PDF/ UA Reults", "WCAG Reults", "Adobe Acrobat Audit Status"},
    ),
    (
        "WCAG",
        [
            "PDF Name",
            "Original URL",
            "Final URL",
            "WCAG Reults",
            "WCAG Notes",
            "WCAG Criteria Affected",
            "WCAG Adobe Findings",
            "WCAG Adobe Statuses",
            "WCAG Fix Directions",
            "WCAG Machine-Verifiable",
            "Colour contrast machine check",
            "Reading order machine check",
            "Adobe accessibility check",
        ],
        {"WCAG Reults", "Colour contrast machine check", "Reading order machine check", "Adobe accessibility check"},
    ),
    (
        "PDF UA + Semantics",
        [
            "PDF Name",
            "Original URL",
            "Final URL",
            "PDF/ UA Reults",
            "PDF/UA Notes",
            "Document title present",
            "Document language present",
            "Assistive technology access not blocked",
            "Extractable text present",
            "Tagged PDF present",
            "At least one heading present",
            "Figure alt text present",
            "Bookmarks present when required",
            "TOC present for general documents",
            "Lists use L tag",
            "Tables use Table tag",
            "Links use Link tag",
            "Form fields have tooltips",
            "Footnotes use Note tag",
        ],
        {
            "PDF/ UA Reults",
            "Document title present",
            "Document language present",
            "Assistive technology access not blocked",
            "Extractable text present",
            "Tagged PDF present",
            "At least one heading present",
            "Figure alt text present",
            "Bookmarks present when required",
            "TOC present for general documents",
            "Lists use L tag",
            "Tables use Table tag",
            "Links use Link tag",
            "Form fields have tooltips",
            "Footnotes use Note tag",
        },
    ),
    (
        "Remediation",
        [
            "PDF Name",
            "Original URL",
            "Final URL",
            "Overall Result",
            "PDF/ UA Reults",
            "WCAG Reults",
            "Failure Summary",
            "Failure Detail",
            "Remediation Guidance",
            "Manual Review Summary",
            "Remediation Notes",
            "WCAG Criteria Affected",
            "WCAG Fix Directions",
        ],
        {"Overall Result", "PDF/ UA Reults", "WCAG Reults"},
    ),
    (
        "Source Retrieval",
        [
            "PDF Name",
            "Original URL",
            "Final URL",
            "Retrieval Category",
            "HTTP Status",
            "Redirect Count",
            "Content Type",
            "Page Title",
            "Page Count",
            "URL resolves",
            "URL is expected file",
            "Unexpected landing page",
            "Notes",
        ],
        {"Retrieval Category", "URL resolves", "URL is expected file", "Unexpected landing page"},
    ),
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="E8EEF5")
HEADER_FONT = Font(bold=True, color="1F2937")
THIN_BORDER = Border(
    left=Side(style="thin", color="FFFFFF"),
    right=Side(style="thin", color="FFFFFF"),
    top=Side(style="thin", color="FFFFFF"),
    bottom=Side(style="thin", color="FFFFFF"),
)


def load_flat_records(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(i for i, row in enumerate(rows) if row and "PDF Name" in row)
    header = [str(value) if value is not None else "" for value in rows[header_idx]]
    records: list[dict[str, object]] = []
    for row in rows[header_idx + 1 :]:
        if not any(value is not None and value != "" for value in row):
            continue
        records.append({header[i]: row[i] if i < len(row) else None for i in range(len(header))})
    return header, records


def apply_sheet_formatting(ws, status_columns: set[str]) -> None:  # type: ignore[no-untyped-def]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        max_length = len(str(cell.value or ""))
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            text = str(value)
            max_length = max(max_length, len(text))
            body_cell = ws.cell(row=row_idx, column=col_idx)
            body_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if cell.value in status_columns:
                style = STATUS_STYLE_MAP.get(text)
                if style:
                    body_cell.fill = PatternFill(fill_type="solid", fgColor=style["fill"])
                    body_cell.font = Font(color=style["font"], bold=True)
                    body_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    body_cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 14), 42)


def build_audit_report(source_path: Path, output_path: Path) -> Path:
    _, records = load_flat_records(source_path)
    out_wb = Workbook()
    default = out_wb.active
    out_wb.remove(default)

    for sheet_name, columns, status_columns in SHEET_SPECS:
        ws = out_wb.create_sheet(title=sheet_name)
        ws.append(columns)
        url_cols = {"Original URL", "Final URL"}
        for row_idx, record in enumerate(records, start=2):
            ws.append([record.get(column) for column in columns])
            for col_idx, column in enumerate(columns, start=1):
                if column in url_cols:
                    value = ws.cell(row=row_idx, column=col_idx).value
                    if isinstance(value, str) and value.startswith("http"):
                        ws.cell(row=row_idx, column=col_idx).hyperlink = value
                if column == "PDF Name":
                    url = record.get("Original URL")
                    if isinstance(url, str) and url.startswith("http"):
                        ws.cell(row=row_idx, column=col_idx).hyperlink = url
        apply_sheet_formatting(ws, status_columns)

    out_wb.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a multi-sheet workbook from the flat audit export.")
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    build_audit_report(args.source, args.output)
    print(args.output)


if __name__ == "__main__":
    main()
