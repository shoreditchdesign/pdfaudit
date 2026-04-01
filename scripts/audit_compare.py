from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def _normalize_bucket(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def main(control_json_path: str, report_xlsx_path: str, output_csv_path: str) -> int:
    control_rows = json.loads(Path(control_json_path).read_text())
    workbook = load_workbook(Path(report_xlsx_path), data_only=True)
    worksheet = workbook["Sheet1"]
    header_row_index = 1
    first_row = [cell.value for cell in worksheet[1]]
    if "Original URL" not in first_row and worksheet.max_row >= 2:
        second_row = [cell.value for cell in worksheet[2]]
        if "Original URL" in second_row:
            header_row_index = 2
            headers = second_row
        else:
            headers = first_row
    else:
        headers = first_row
    header_index = {str(name): idx + 1 for idx, name in enumerate(headers) if name is not None}

    report_rows: dict[str, dict[str, object]] = {}

    def cell_value(row_idx: int, column_name: str) -> object:
        col_idx = header_index.get(column_name)
        if col_idx is None:
            return ""
        return worksheet.cell(row_idx, col_idx).value

    for row_idx in range(header_row_index + 1, worksheet.max_row + 1):
        original_url = cell_value(row_idx, "Original URL")
        if not original_url:
            continue
        report_rows[str(original_url)] = {
            "pdf_name": cell_value(row_idx, "PDF Name"),
            "axes_status": cell_value(row_idx, "Axes Audit Status"),
            "pdfua": cell_value(row_idx, "PDF/ UA Reults"),
            "wcag": cell_value(row_idx, "WCAG Reults"),
            "pdfua_notes": cell_value(row_idx, "PDF/UA Notes"),
            "wcag_notes": cell_value(row_idx, "WCAG Notes"),
            "adobe_status": cell_value(row_idx, "Adobe Acrobat Audit Status"),
            "final_url": cell_value(row_idx, "Final URL"),
            "retrieval_category": cell_value(row_idx, "Retrieval Category"),
            "overall_result": cell_value(row_idx, "Overall Result"),
            "failure_summary": cell_value(row_idx, "Failure Summary"),
        }

    output_path = Path(output_csv_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "Source Row Number",
                "PDF Name",
                "URL",
                "Control Axes Status",
                "Our Axes Status",
                "Axes Match",
                "Control PDF/UA",
                "Our PDF/UA",
                "PDF/UA Match",
                "Control WCAG",
                "Our WCAG",
                "WCAG Match",
                "Final URL",
                "Retrieval Category",
                "Overall Result",
                "Failure Summary",
                "Control PDF/UA Notes",
                "Our PDF/UA Notes",
                "Control WCAG Notes",
                "Our WCAG Notes",
                "Adobe Acrobat Audit Status",
            ],
        )
        writer.writeheader()

        for item in control_rows:
            report_row = report_rows.get(item["url"], {})
            control_axes = _normalize_bucket(item.get("control_axes_status"))
            our_axes = _normalize_bucket(report_row.get("axes_status"))
            control_pdfua = _normalize_bucket(item.get("control_pdfua"))
            our_pdfua = _normalize_bucket(report_row.get("pdfua"))
            control_wcag = _normalize_bucket(item.get("control_wcag"))
            our_wcag = _normalize_bucket(report_row.get("wcag"))
            writer.writerow(
                {
                    "Source Row Number": item.get("source_row_number"),
                    "PDF Name": item.get("pdf_name"),
                    "URL": item.get("url"),
                    "Control Axes Status": control_axes,
                    "Our Axes Status": our_axes,
                    "Axes Match": str(bool(control_axes and control_axes == our_axes)),
                    "Control PDF/UA": control_pdfua,
                    "Our PDF/UA": our_pdfua,
                    "PDF/UA Match": str(bool(control_pdfua and control_pdfua == our_pdfua)),
                    "Control WCAG": control_wcag,
                    "Our WCAG": our_wcag,
                    "WCAG Match": str(bool(control_wcag and control_wcag == our_wcag)),
                    "Final URL": report_row.get("final_url"),
                    "Retrieval Category": report_row.get("retrieval_category"),
                    "Overall Result": report_row.get("overall_result"),
                    "Failure Summary": report_row.get("failure_summary"),
                    "Control PDF/UA Notes": item.get("control_pdfua_notes"),
                    "Our PDF/UA Notes": report_row.get("pdfua_notes"),
                    "Control WCAG Notes": item.get("control_wcag_notes"),
                    "Our WCAG Notes": report_row.get("wcag_notes"),
                    "Adobe Acrobat Audit Status": report_row.get("adobe_status"),
                }
            )

    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1], sys.argv[2], sys.argv[3]))
