from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
import re
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.domain import CheckStatus, DocumentAuditRecord, RetrievalCategory, SummaryMetrics
from app.services.adobe_cache_service import AdobeReportStore
from app.services.audit_rule_catalog import RULE_DEFINITIONS


TRACKER_COLUMNS = [
    "PDF Name",
    "Total Traffic",
    "Date Audited",
    "PDF/ UA Reults",
    "WCAG Reults",
    "PDF/UA Notes",
    "WCAG Notes",
    "Remediation Status",
    "Adobe Acrobat Audit Status",
    "Remediation Notes",
]

WCAG_DETAIL_COLUMNS = [
    "WCAG Criteria Affected",
    "WCAG Adobe Findings",
    "WCAG Adobe Statuses",
    "WCAG Fix Directions",
    "WCAG Machine-Verifiable",
]

TECHNICAL_COLUMNS = [
    "Original URL",
    "Final URL",
    "Retrieval Category",
    "HTTP Status",
    "Redirect Count",
    "Content Type",
    "Page Title",
    "Page Count",
    "Overall Result",
    "HSBC Policy Result",
    "Failure Summary",
    "Failure Detail",
    "Remediation Guidance",
    "Manual Review Summary",
    "Notes",
]

RULE_COLUMNS = [
    "url_resolves",
    "url_is_expected_file",
    "unexpected_landing_page",
    "doc_title_present",
    "doc_language_present",
    "at_access_not_blocked",
    "extractable_text_present",
    "tagged_pdf_present",
    "at_least_one_heading",
    "figure_alt_present",
    "bookmarks_present_if_gt_3_pages",
    "toc_present_general_doc_if_ge_5",
    "lists_use_l",
    "tables_use_table_tag",
    "links_use_link_tag",
    "form_fields_have_tu",
    "footnotes_use_note",
    "colour_contrast_machine_check",
    "reading_order_machine_check",
    "adobe_full_check",
]

REPORT_COLUMNS = [
    *TRACKER_COLUMNS,
    *WCAG_DETAIL_COLUMNS,
    *TECHNICAL_COLUMNS,
    *[RULE_DEFINITIONS[rule_id].title for rule_id in RULE_COLUMNS],
]

HEADER_GROUPS = [
    ("Tracker Summary", TRACKER_COLUMNS),
    ("WCAG Detail", WCAG_DETAIL_COLUMNS),
    ("Source & Retrieval", TECHNICAL_COLUMNS[:8]),
    ("Audit Summary", TECHNICAL_COLUMNS[8:12]),
    ("Remediation & Detail", TECHNICAL_COLUMNS[12:15]),
    ("Retrieval Checks", [RULE_DEFINITIONS[rule_id].title for rule_id in RULE_COLUMNS[:3]]),
    (
        "Core PDF/UA",
        [
            RULE_DEFINITIONS["doc_title_present"].title,
            RULE_DEFINITIONS["doc_language_present"].title,
            RULE_DEFINITIONS["at_access_not_blocked"].title,
            RULE_DEFINITIONS["extractable_text_present"].title,
            RULE_DEFINITIONS["tagged_pdf_present"].title,
        ],
    ),
    (
        "Semantic Review",
        [
            RULE_DEFINITIONS["at_least_one_heading"].title,
            RULE_DEFINITIONS["figure_alt_present"].title,
            RULE_DEFINITIONS["bookmarks_present_if_gt_3_pages"].title,
            RULE_DEFINITIONS["toc_present_general_doc_if_ge_5"].title,
            RULE_DEFINITIONS["lists_use_l"].title,
            RULE_DEFINITIONS["tables_use_table_tag"].title,
            RULE_DEFINITIONS["links_use_link_tag"].title,
            RULE_DEFINITIONS["form_fields_have_tu"].title,
            RULE_DEFINITIONS["footnotes_use_note"].title,
        ],
    ),
    (
        "Adobe Checks",
        [
            RULE_DEFINITIONS["colour_contrast_machine_check"].title,
            RULE_DEFINITIONS["reading_order_machine_check"].title,
            RULE_DEFINITIONS["adobe_full_check"].title,
        ],
    ),
]

GROUP_FILLS = {
    "Tracker Summary": "1F4E78",
    "Source & Retrieval": "245C69",
    "Audit Summary": "4F6D3A",
    "Remediation & Detail": "6E5C2F",
    "Retrieval Checks": "375A7F",
    "Core PDF/UA": "1B6E65",
    "Semantic Review": "5E7A2F",
    "Adobe Checks": "7F3F1B",
}

ILLEGAL_XLSX_CHAR_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

STATUS_STYLE_MAP = {
    "PASS": {"fill": "DCFCE7", "font": "166534"},
    "Pass": {"fill": "DCFCE7", "font": "166534"},
    "Needs Review": {"fill": "FEF3C7", "font": "92400E"},
    "Pending Adobe": {"fill": "E5E7EB", "font": "4B5563"},
    "COMPLETE": {"fill": "DBEAFE", "font": "1D4ED8"},
    "Complete": {"fill": "DBEAFE", "font": "1D4ED8"},
    "FAIL": {"fill": "FEE2E2", "font": "B91C1C"},
    "Fail": {"fill": "FEE2E2", "font": "B91C1C"},
    "HARD_404": {"fill": "FEE2E2", "font": "B91C1C"},
    "SOFT_404": {"fill": "FEE2E2", "font": "B91C1C"},
    "REQUEST_ERROR": {"fill": "FEE2E2", "font": "B91C1C"},
    "BLOCKED": {"fill": "FDE68A", "font": "92400E"},
    "Blocked": {"fill": "FDE68A", "font": "92400E"},
    "ALT_LANDING_PAGE_NON_PDF": {"fill": "FDE68A", "font": "92400E"},
    "REVIEW_REQUIRED": {"fill": "FEF3C7", "font": "92400E"},
    "NEEDS_MANUAL_REVIEW": {"fill": "FEF3C7", "font": "92400E"},
    "API_UNAVAILABLE": {"fill": "E5E7EB", "font": "4B5563"},
    "N/A": {"fill": "F3F4F6", "font": "6B7280"},
    "Not Started": {"fill": "F3F4F6", "font": "6B7280"},
}

TRACKER_STATUS_COLUMNS = {
    "Axes Audit Status",
    "PDF/ UA Reults",
    "WCAG Reults",
    "Remediation Status",
    "Adobe Acrobat Audit Status",
}

TECHNICAL_STATUS_COLUMNS = {
    "Retrieval Category",
    "Overall Result",
    "HSBC Policy Result",
}

WRAP_TEXT_COLUMNS = {
    *TRACKER_COLUMNS,
    *WCAG_DETAIL_COLUMNS,
    *TECHNICAL_COLUMNS,
    *[RULE_DEFINITIONS[rule_id].title for rule_id in RULE_COLUMNS],
}

PREFERRED_COLUMN_WIDTHS = {
    "PDF Name": 28,
    "PDF/UA Notes": 42,
    "WCAG Notes": 42,
    "Remediation Notes": 42,
    "WCAG Criteria Affected": 34,
    "WCAG Adobe Findings": 28,
    "WCAG Adobe Statuses": 28,
    "WCAG Fix Directions": 42,
    "WCAG Machine-Verifiable": 26,
    "Original URL": 24,
    "Final URL": 24,
    "Page Title": 28,
    "Failure Summary": 34,
    "Failure Detail": 56,
    "Remediation Guidance": 44,
    "Manual Review Summary": 44,
    "Notes": 44,
}

THIN_BORDER = Border(
    left=Side(style="thin", color="FFFFFF"),
    right=Side(style="thin", color="FFFFFF"),
    top=Side(style="thin", color="FFFFFF"),
    bottom=Side(style="thin", color="FFFFFF"),
)

MULTISHEET_SPECS = [
    (
        "Summary",
        [
            "PDF Name",
            "Original URL",
            "Final URL",
            "Date Audited",
            "Overall Result",
            "PDF/ UA Reults",
            "WCAG Reults",
            "Adobe Acrobat Audit Status",
            "PDF/UA Notes",
            "WCAG Notes",
            "Failure Summary",
        ],
        {"Overall Result", "PDF/ UA Reults", "WCAG Reults", "Adobe Acrobat Audit Status"},
    ),
    (
        "WCAG",
        [
            "PDF Name",
            "Original URL",
            "Final URL",
            "WCAG Reults",
            "WCAG Notes",
            "WCAG Criteria Affected",
            "WCAG Adobe Findings",
            "WCAG Adobe Statuses",
            "WCAG Fix Directions",
            "WCAG Machine-Verifiable",
            "Colour contrast machine check",
            "Reading order machine check",
            "Adobe accessibility check",
        ],
        {"WCAG Reults", "Colour contrast machine check", "Reading order machine check", "Adobe accessibility check"},
    ),
    (
        "PDF UA + Semantics",
        [
            "PDF Name",
            "Original URL",
            "Final URL",
            "PDF/ UA Reults",
            "PDF/UA Notes",
            "Document title present",
            "Document language present",
            "Assistive technology access not blocked",
            "Extractable text present",
            "Tagged PDF present",
            "At least one heading present",
            "Figure alt text present",
            "Bookmarks present when required",
            "TOC present for general documents",
            "Lists use L tag",
            "Tables use Table tag",
            "Links use Link tag",
            "Form fields have tooltips",
            "Footnotes use Note tag",
        ],
        {
            "PDF/ UA Reults",
            "Document title present",
            "Document language present",
            "Assistive technology access not blocked",
            "Extractable text present",
            "Tagged PDF present",
            "At least one heading present",
            "Figure alt text present",
            "Bookmarks present when required",
            "TOC present for general documents",
            "Lists use L tag",
            "Tables use Table tag",
            "Links use Link tag",
            "Form fields have tooltips",
            "Footnotes use Note tag",
        },
    ),
    (
        "Remediation",
        [
            "PDF Name",
            "Original URL",
            "Final URL",
            "Overall Result",
            "PDF/ UA Reults",
            "WCAG Reults",
            "Failure Summary",
            "Failure Detail",
            "Remediation Guidance",
            "Manual Review Summary",
            "Remediation Notes",
            "WCAG Criteria Affected",
            "WCAG Fix Directions",
        ],
        {"Overall Result", "PDF/ UA Reults", "WCAG Reults"},
    ),
    (
        "Source Retrieval",
        [
            "PDF Name",
            "Original URL",
            "Final URL",
            "Retrieval Category",
            "HTTP Status",
            "Redirect Count",
            "Content Type",
            "Page Title",
            "Page Count",
            "URL resolves",
            "URL is expected file",
            "Unexpected landing page",
            "Notes",
        ],
        {"Retrieval Category", "URL resolves", "URL is expected file", "Unexpected landing page"},
    ),
]

ADOBE_WCAG_RULE_MAP = {
    "color contrast": (
        "WCAG 1.4.3 Contrast (Minimum)",
        "Review the foreground/background color pairings Adobe flagged and correct any text that does not meet AA contrast ratios.",
    ),
    "logical reading order": (
        "WCAG 1.3.2 Meaningful Sequence",
        "Confirm the reading order is meaningful for assistive technology and repair the source structure or tag order where needed.",
    ),
    "figures alternate text": (
        "WCAG 1.1.1 Non-text Content",
        "Add or correct alternate text for non-decorative figures so the same purpose or meaning is available non-visually.",
    ),
    "field descriptions": (
        "WCAG 3.3.2 Labels or Instructions / 4.1.2 Name, Role, Value",
        "Add clear programmatic descriptions to form fields so users receive accurate labels and instructions.",
    ),
    "tagged content": (
        "WCAG 1.3.1 Info and Relationships",
        "Repair the tag structure so visible content is represented semantically for assistive technology.",
    ),
    "tagged annotations": (
        "WCAG 4.1.2 Name, Role, Value",
        "Ensure annotations and interactive elements are tagged so assistive technology can expose their purpose correctly.",
    ),
    "headers": (
        "WCAG 1.3.1 Info and Relationships",
        "Add proper table headers or equivalent structural markup so relationships inside tables are programmatically available.",
    ),
    "summary": (
        "WCAG 1.3.1 Info and Relationships",
        "Add a meaningful table summary or equivalent contextual description where Adobe flagged missing table summaries.",
    ),
    "title": (
        "WCAG 2.4.2 Page Titled",
        "Set a meaningful document title and ensure the viewer shows it instead of the filename.",
    ),
    "primary language": (
        "WCAG 3.1.1 Language of Page",
        "Set the document primary language in metadata so assistive technology uses the correct pronunciation rules.",
    ),
    "bookmarks": (
        "WCAG 2.4.1 Bypass Blocks / 2.4.5 Multiple Ways",
        "Add bookmarks for longer documents so users can move between sections more efficiently.",
    ),
    "tagged form fields": (
        "WCAG 1.3.1 Info and Relationships / 4.1.2 Name, Role, Value",
        "Tag all form fields so their role and relationship to the surrounding structure are available programmatically.",
    ),
    "tab order": (
        "WCAG 2.4.3 Focus Order",
        "Align tab order with the document structure so keyboard users move through the PDF in a logical order.",
    ),
    "character encoding": (
        "WCAG 4.1.2 Name, Role, Value",
        "Repair text encoding issues so assistive technology can reliably interpret the document content.",
    ),
}


class WorkbookTemplateResolver:
    def __init__(self, template_path: Path | None) -> None:
        self.template_path = template_path

    def create_workbook(self) -> Workbook:
        if self.template_path and self.template_path.exists():
            workbook = load_workbook(self.template_path)
            while len(workbook.worksheets) > 1:
                workbook.remove(workbook.worksheets[-1])
            sheet = workbook.worksheets[0]
            sheet.title = "Sheet1"
            if sheet.max_row > 0:
                sheet.delete_rows(1, sheet.max_row)
            return workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        return workbook


class ReportBuilder:
    def __init__(
        self,
        template_resolver: WorkbookTemplateResolver,
        adobe_report_store: AdobeReportStore | None = None,
    ) -> None:
        self.template_resolver = template_resolver
        self.adobe_report_store = adobe_report_store

    def build(self, rows: list[DocumentAuditRecord], summary: SummaryMetrics) -> bytes:
        workbook = self.template_resolver.create_workbook()
        self._reset_workbook(workbook)

        flat_rows = [self._flat_record(row, summary) for row in rows]
        for sheet_name, columns, status_columns in MULTISHEET_SPECS:
            sheet = workbook.create_sheet(title=sheet_name)
            sheet.append(columns)
            self._style_simple_header(sheet)
            for row_index, flat_row in enumerate(flat_rows, start=2):
                sheet.append([self._clean_cell_value(flat_row.get(column)) for column in columns])
                for col_index, column in enumerate(columns, start=1):
                    cell = sheet.cell(row=row_index, column=col_index)
                    if column in {"Original URL", "Final URL"}:
                        if isinstance(cell.value, str) and cell.value.startswith("http"):
                            cell.hyperlink = cell.value
                    elif column == "PDF Name":
                        url = flat_row.get("Original URL")
                        if isinstance(url, str) and url.startswith("http"):
                            cell.hyperlink = url
                self._apply_multisheet_status_styles(sheet, row_index, columns, status_columns)
            self._format_multisheet(sheet, columns)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()

    @staticmethod
    def _reset_workbook(workbook: Workbook) -> None:
        for sheet in list(workbook.worksheets):
            workbook.remove(sheet)

    def _flat_record(self, row: DocumentAuditRecord, summary: SummaryMetrics) -> dict[str, object]:
        tracker_row = self._tracker_row(row, summary)
        wcag_detail_row = self._wcag_detail_row(row)
        technical_row = self._technical_row(row)
        rule_row = [row.rule_results[rule_id].status.value for rule_id in RULE_COLUMNS]
        values = [*tracker_row, *wcag_detail_row, *technical_row, *rule_row]
        return {REPORT_COLUMNS[index]: values[index] for index in range(len(REPORT_COLUMNS))}

    @staticmethod
    def _style_simple_header(sheet) -> None:  # type: ignore[no-untyped-def]
        child_font = Font(bold=True, color="1F2937")
        child_fill = PatternFill(fill_type="solid", fgColor="E8EEF5")
        for cell in sheet[1]:
            cell.font = child_font
            cell.fill = child_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[1].height = 28

    def _format_multisheet(self, sheet, columns: list[str]) -> None:  # type: ignore[no-untyped-def]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        for idx, column_name in enumerate(columns, start=1):
            max_length = len(str(column_name))
            for row in range(2, sheet.max_row + 1):
                value = sheet.cell(row=row, column=idx).value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
                sheet.cell(row=row, column=idx).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            preferred_width = PREFERRED_COLUMN_WIDTHS.get(column_name)
            width = preferred_width if preferred_width is not None else min(max(max_length + 2, 12), 42)
            sheet.column_dimensions[get_column_letter(idx)].width = width

    def _apply_multisheet_status_styles(
        self,
        sheet,
        row_index: int,
        columns: list[str],
        status_columns: set[str],
    ) -> None:  # type: ignore[no-untyped-def]
        for col_index, column_name in enumerate(columns, start=1):
            if column_name not in status_columns:
                continue
            cell = sheet.cell(row=row_index, column=col_index)
            if cell.value is None:
                continue
            style = STATUS_STYLE_MAP.get(str(cell.value))
            if not style:
                continue
            cell.fill = PatternFill(fill_type="solid", fgColor=style["fill"])
            cell.font = Font(color=style["font"], bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    @staticmethod
    def _group_header_row() -> list[str]:
        group_row: list[str] = []
        for group_name, columns in HEADER_GROUPS:
            group_row.append(group_name)
            group_row.extend([""] * (len(columns) - 1))
        return group_row

    def _style_headers(self, sheet) -> None:  # type: ignore[no-untyped-def]
        group_font = Font(color="FFFFFF", bold=True)
        child_font = Font(bold=True, color="1F2937")
        child_fill = PatternFill(fill_type="solid", fgColor="E8EEF5")
        for start_col, (group_name, columns) in self._group_spans().items():
            end_col = start_col + len(columns) - 1
            sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = sheet.cell(row=1, column=start_col)
            cell.fill = PatternFill(fill_type="solid", fgColor=GROUP_FILLS.get(group_name, "334155"))
            cell.font = group_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

            for col in range(start_col, end_col + 1):
                child = sheet.cell(row=2, column=col)
                child.font = child_font
                child.fill = child_fill
                child.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        sheet.row_dimensions[1].height = 24
        sheet.row_dimensions[2].height = 36

    def _format_sheet(self, sheet) -> None:  # type: ignore[no-untyped-def]
        sheet.freeze_panes = "A3"
        sheet.auto_filter.ref = f"A2:{get_column_letter(sheet.max_column)}{sheet.max_row}"

        for idx, column_name in enumerate(REPORT_COLUMNS, start=1):
            max_length = len(str(column_name))
            for row in range(3, sheet.max_row + 1):
                value = sheet.cell(row=row, column=idx).value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
            preferred_width = PREFERRED_COLUMN_WIDTHS.get(column_name)
            width = preferred_width if preferred_width is not None else min(max(max_length + 2, 12), 42)
            sheet.column_dimensions[get_column_letter(idx)].width = width

    def _apply_status_styles(self, sheet) -> None:  # type: ignore[no-untyped-def]
        status_columns = {
            idx
            for idx, name in enumerate(REPORT_COLUMNS, start=1)
            if name in TRACKER_STATUS_COLUMNS or name in TECHNICAL_STATUS_COLUMNS or name in {
                RULE_DEFINITIONS[rule_id].title for rule_id in RULE_COLUMNS
            }
        }

        for row in range(3, sheet.max_row + 1):
            for col in status_columns:
                cell = sheet.cell(row=row, column=col)
                value = cell.value
                if value is None:
                    continue
                style = STATUS_STYLE_MAP.get(str(value))
                if not style:
                    continue
                cell.fill = PatternFill(fill_type="solid", fgColor=style["fill"])
                cell.font = Font(color=style["font"], bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = THIN_BORDER

    def _apply_body_alignment(self, sheet) -> None:  # type: ignore[no-untyped-def]
        for idx, column_name in enumerate(REPORT_COLUMNS, start=1):
            if column_name not in WRAP_TEXT_COLUMNS:
                continue
            for row in range(3, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=idx)
                if cell.value is None:
                    continue
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    @staticmethod
    def _group_spans() -> dict[int, tuple[str, list[str]]]:
        spans: dict[int, tuple[str, list[str]]] = {}
        cursor = 1
        for group_name, columns in HEADER_GROUPS:
            spans[cursor] = (group_name, columns)
            cursor += len(columns)
        return spans

    def _tracker_row(self, row: DocumentAuditRecord, summary: SummaryMetrics) -> list[object]:
        return [
            self._pdf_name(row),
            None,
            summary.run_timestamp.date(),
            self._tracker_bucket_value(row.grouped_results.pdf_ua_result),
            self._tracker_bucket_value(row.grouped_results.wcag_result),
            self._pdf_ua_notes(row),
            self._wcag_notes(row),
            "Not Started",
            self._adobe_status(row),
            self._remediation_notes(row),
        ]

    def _technical_row(self, row: DocumentAuditRecord) -> list[object]:
        return [
            row.original_url,
            row.final_url,
            row.retrieval_category.value,
            row.http_status,
            row.redirect_count,
            row.content_type,
            row.page_title,
            row.page_count,
            row.overall_result.value,
            row.grouped_results.hsbc_policy_result.value,
            row.failure_summary,
            row.failure_detail,
            row.remediation_guidance,
            row.manual_review_summary,
            row.notes,
        ]

    def _wcag_detail_row(self, row: DocumentAuditRecord) -> list[object]:
        findings = self._mapped_wcag_findings(row)
        if not findings:
            return [None, None, None, None, None]
        criteria = " | ".join(self._unique_in_order(item["criterion"] for item in findings))
        adobe_findings = " | ".join(self._unique_in_order(item["rule"] for item in findings))
        adobe_statuses = " | ".join(
            self._unique_in_order(f'{item["rule"]}: {item["status"]}' for item in findings)
        )
        fix_directions = " | ".join(self._unique_in_order(item["fix"] for item in findings))
        machine_verifiable = " | ".join(
            self._unique_in_order(
                f'{item["rule"]}: {"No" if item["status"] == "Needs manual check" else "Yes"}' for item in findings
            )
        )
        return [criteria, adobe_findings, adobe_statuses, fix_directions, machine_verifiable]

    @staticmethod
    def _pdf_name(row: DocumentAuditRecord) -> str:
        if row.pdf_name:
            return row.pdf_name
        filename = Path(urlparse(row.original_url).path).stem
        return filename.replace("-", " ").replace("_", " ").title()

    @staticmethod
    def _tracker_bucket_value(status: CheckStatus) -> str | None:
        if status == CheckStatus.PASS:
            return "Pass"
        if status == CheckStatus.FAIL:
            return "Fail"
        if status == CheckStatus.NEEDS_MANUAL_REVIEW:
            return "Needs Review"
        if status == CheckStatus.API_UNAVAILABLE:
            return "Pending Adobe"
        return None

    def _pdf_ua_notes(self, row: DocumentAuditRecord) -> str | None:
        if row.retrieval_category != RetrievalCategory.DIRECT_FILE_OK:
            if row.retrieval_category == RetrievalCategory.REVIEW_REQUIRED:
                return "The linked asset is not a PDF. Route this row to the correct Office-file review path instead of PDF/UA testing."
            return "The source PDF could not be retrieved for accessibility testing."

        if self._has_tagged_pdf_failure(row):
            return (
                "Basic requirement for accessible PDFs not met!\n\n"
                "The PDF does not contain an invisible structural layer in the form of tags and "
                "therefore does not meet the requirements for further accessibility testing."
            )

        if row.grouped_results.pdf_ua_result == CheckStatus.PASS:
            return "The file meets all machine-verifiable PDF/UA requirements."

        findings = self._bucket_findings(
            row,
            {
                "doc_title_present",
                "doc_language_present",
                "at_access_not_blocked",
                "tagged_pdf_present",
                "figure_alt_present",
                "bookmarks_present_if_gt_3_pages",
                "lists_use_l",
                "tables_use_table_tag",
                "links_use_link_tag",
                "form_fields_have_tu",
                "footnotes_use_note",
            },
            exclude_adobe_manual_only=True,
        )
        if findings:
            return " | ".join(findings[:4])
        return None

    def _wcag_notes(self, row: DocumentAuditRecord) -> str | None:
        if row.retrieval_category != RetrievalCategory.DIRECT_FILE_OK:
            if row.retrieval_category == RetrievalCategory.REVIEW_REQUIRED:
                return "The linked asset is an Office-format file rather than a PDF, so PDF WCAG results do not apply until it is converted or reviewed in the correct format."
            return "The source PDF resolved to a not-found or non-PDF destination."

        if row.grouped_results.wcag_result == CheckStatus.API_UNAVAILABLE:
            return "Adobe machine checks were unavailable for this file. Re-run the Adobe step before treating WCAG as final."

        if row.grouped_results.wcag_result == CheckStatus.NEEDS_MANUAL_REVIEW:
            return "Adobe flagged items that need human review before a final WCAG verdict is assigned."

        if row.grouped_results.wcag_result == CheckStatus.PASS:
            return "The file meets all machine-verifiable WCAG requirements (Level A and AA)."

        if self._has_adobe_rule_status(row, "Color contrast", CheckStatus.NEEDS_MANUAL_REVIEW):
            return (
                "Adobe flagged color contrast for manual review. This typically indicates text/background "
                "contrast may not meet accessibility expectations."
            )

        findings = self._bucket_findings(
            row,
            {
                "url_resolves",
                "url_is_expected_file",
                "extractable_text_present",
                "colour_contrast_machine_check",
                "reading_order_machine_check",
                "adobe_full_check",
            },
            exclude_adobe_manual_only=False,
        )
        if findings:
            return " | ".join(findings[:4])
        return None

    def _adobe_status(self, row: DocumentAuditRecord) -> str:
        adobe_result = row.rule_results["adobe_full_check"].status
        if adobe_result in {CheckStatus.NA, CheckStatus.API_UNAVAILABLE}:
            return "Not Started"
        return "Complete"

    def _remediation_notes(self, row: DocumentAuditRecord) -> str | None:
        parts = [
            part
            for part in [
                row.remediation_guidance,
                row.manual_review_summary,
                self._wcag_remediation_notes_from_adobe(row),
            ]
            if part
        ]
        return " | ".join(parts) if parts else None

    @staticmethod
    def _clean_cell_value(value: object) -> object:
        if isinstance(value, str):
            return ILLEGAL_XLSX_CHAR_RE.sub("", value)
        return value

    def _adobe_findings(self, row: DocumentAuditRecord) -> list[dict]:
        raw = self._adobe_raw_report(row)
        detailed = raw.get("Detailed Report")
        if not isinstance(detailed, dict):
            return []
        findings: list[dict] = []
        for section_name, group in detailed.items():
            if isinstance(group, list):
                for finding in group:
                    if isinstance(finding, dict):
                        enriched = dict(finding)
                        enriched["_section"] = section_name
                        findings.append(enriched)
        return findings

    def _adobe_raw_report(self, row: DocumentAuditRecord) -> dict:
        raw = row.rule_results["adobe_full_check"].raw
        if isinstance(raw, dict) and raw:
            return raw
        if self.adobe_report_store:
            cached = self.adobe_report_store.load(row.original_url) or self.adobe_report_store.load(row.final_url)
            if cached:
                return cached
        return {}

    def _has_adobe_rule_failure(self, row: DocumentAuditRecord, rule_name: str) -> bool:
        return self._has_adobe_rule_status(row, rule_name, CheckStatus.FAIL)

    def _has_tagged_pdf_failure(self, row: DocumentAuditRecord) -> bool:
        tagged_pdf_result = row.rule_results.get("tagged_pdf_present")
        if tagged_pdf_result and tagged_pdf_result.status == CheckStatus.FAIL:
            return True
        return self._has_adobe_rule_failure(row, "Tagged PDF")

    def _has_adobe_rule_status(self, row: DocumentAuditRecord, rule_name: str, status: CheckStatus) -> bool:
        for finding in self._adobe_findings(row):
            rule = str(finding.get("Rule", "")).strip().lower()
            finding_status = str(finding.get("Status", "")).strip().lower()
            if rule == rule_name.lower():
                if status == CheckStatus.FAIL and finding_status == "failed":
                    return True
                if status == CheckStatus.NEEDS_MANUAL_REVIEW and finding_status == "needs manual check":
                    return True
        return False

    def _bucket_findings(
        self,
        row: DocumentAuditRecord,
        rule_ids: set[str],
        *,
        exclude_adobe_manual_only: bool,
    ) -> list[str]:
        findings: list[str] = []
        for rule_id in rule_ids:
            result = row.rule_results.get(rule_id)
            if not result:
                continue
            if result.status not in {CheckStatus.FAIL, CheckStatus.NEEDS_MANUAL_REVIEW}:
                continue
            if exclude_adobe_manual_only and rule_id in {"colour_contrast_machine_check", "reading_order_machine_check"}:
                continue
            title = RULE_DEFINITIONS[rule_id].title
            detail = result.evidence[0] if result.evidence else title
            findings.append(detail if detail != title else f"{title} needs attention.")
        return findings

    def _wcag_remediation_notes_from_adobe(self, row: DocumentAuditRecord) -> str | None:
        mapped_notes: list[str] = []
        seen: set[str] = set()
        for finding in self._mapped_wcag_findings(row):
            criterion = finding["criterion"]
            note = finding["fix"]
            rendered = f"{criterion}: {note}"
            if rendered in seen:
                continue
            seen.add(rendered)
            mapped_notes.append(rendered)
        return " | ".join(mapped_notes) if mapped_notes else None

    def _mapped_wcag_findings(self, row: DocumentAuditRecord) -> list[dict[str, str]]:
        mapped: list[dict[str, str]] = []
        for finding in self._adobe_findings(row):
            status_text = str(finding.get("Status", "")).strip()
            normalized_status = status_text.lower()
            if normalized_status not in {"failed", "needs manual check"}:
                continue
            rule_name = str(finding.get("Rule", "")).strip()
            mapping = ADOBE_WCAG_RULE_MAP.get(rule_name.lower())
            if not mapping:
                continue
            criterion, fix = mapping
            mapped.append(
                {
                    "criterion": criterion,
                    "rule": rule_name,
                    "status": status_text,
                    "fix": fix,
                    "section": str(finding.get("_section", "")).strip(),
                }
            )
        return mapped

    @staticmethod
    def _unique_in_order(items) -> list[str]:
        seen: set[str] = set()
        ordered: list[str] = []
        for item in items:
            value = str(item).strip()
            if not value or value in seen:
                continue
            seen.add(value)
            ordered.append(value)
        return ordered


def compute_summary(rows: list[DocumentAuditRecord]) -> SummaryMetrics:
    summary = SummaryMetrics(run_timestamp=datetime.utcnow())
    summary.total = len(rows)
    for row in rows:
        if row.retrieval_category != RetrievalCategory.DIRECT_FILE_OK:
            summary.unreachable_count += 1

        if row.overall_result == CheckStatus.PASS:
            summary.pass_count += 1
        elif row.overall_result == CheckStatus.FAIL:
            summary.fail_count += 1
        elif row.overall_result == CheckStatus.NEEDS_MANUAL_REVIEW:
            summary.manual_review_count += 1

        for rule_id, result in row.rule_results.items():
            if result.status == CheckStatus.FAIL:
                summary.per_rule_failures[rule_id] = summary.per_rule_failures.get(rule_id, 0) + 1

    return summary
