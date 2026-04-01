from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
import re
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook

from app.models.domain import CheckStatus, DocumentAuditRecord, RetrievalCategory, SummaryMetrics
from app.services.rule_catalog import RULE_DEFINITIONS


TRACKER_COLUMNS = [
    "PDF Name",
    "URL",
    "Category",
    "Total Traffic",
    "Priority Tier",
    "Axes Audit Status",
    "Date Audited",
    "PDF/ UA Reults",
    "WCAG Reults",
    "Link to Audit report",
    "PDF/UA Notes",
    "WCAG Notes",
    "Remediation Status",
    "Adobe Acrobat Audit Status",
    "Remediation Notes",
    "Re-test Result",
    "Link to Audit report",
    "3rd Re-test Result",
    "Link to Audit report",
]

TECHNICAL_COLUMNS = [
    "Original URL",
    "Final URL",
    "Retrieval Category",
    "HTTP Status",
    "Redirect Count",
    "Content Type",
    "Page Title",
    "Page Count",
    "Overall Result",
    "HSBC Policy Result",
    "Failure Summary",
    "Failure Detail",
    "Remediation Guidance",
    "Manual Review Summary",
    "Notes",
]

RULE_COLUMNS = [
    "url_resolves",
    "url_is_expected_file",
    "unexpected_landing_page",
    "doc_title_present",
    "doc_language_present",
    "at_access_not_blocked",
    "extractable_text_present",
    "at_least_one_heading",
    "figure_alt_present",
    "bookmarks_present_if_gt_3_pages",
    "toc_present_general_doc_if_ge_5",
    "lists_use_l",
    "tables_use_table_tag",
    "links_use_link_tag",
    "form_fields_have_tu",
    "footnotes_use_note",
    "colour_contrast_machine_check",
    "reading_order_machine_check",
    "adobe_full_check",
]

REPORT_COLUMNS = [
    *TRACKER_COLUMNS,
    *TECHNICAL_COLUMNS,
    *[RULE_DEFINITIONS[rule_id].title for rule_id in RULE_COLUMNS],
]

ILLEGAL_XLSX_CHAR_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


class WorkbookTemplateResolver:
    def __init__(self, template_path: Path | None) -> None:
        self.template_path = template_path

    def create_workbook(self) -> Workbook:
        if self.template_path and self.template_path.exists():
            workbook = load_workbook(self.template_path)
            while len(workbook.worksheets) > 1:
                workbook.remove(workbook.worksheets[-1])
            sheet = workbook.worksheets[0]
            sheet.title = "Sheet1"
            if sheet.max_row > 0:
                sheet.delete_rows(1, sheet.max_row)
            return workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        return workbook


class ReportBuilder:
    def __init__(self, template_resolver: WorkbookTemplateResolver) -> None:
        self.template_resolver = template_resolver

    def build(self, rows: list[DocumentAuditRecord], summary: SummaryMetrics) -> bytes:
        workbook = self.template_resolver.create_workbook()
        sheet = workbook["Sheet1"]

        if sheet.max_row > 0:
            sheet.delete_rows(1, sheet.max_row)
        sheet.append(REPORT_COLUMNS)

        for row_index, row in enumerate(rows, start=2):
            tracker_row = self._tracker_row(row, summary)
            technical_row = self._technical_row(row)
            rule_row = [row.rule_results[rule_id].status.value for rule_id in RULE_COLUMNS]
            sheet.append([self._clean_cell_value(value) for value in [*tracker_row, *technical_row, *rule_row]])

            url_cell = sheet.cell(row=row_index, column=2)
            url_cell.value = "Open PDF"
            url_cell.hyperlink = row.original_url

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()

    def _tracker_row(self, row: DocumentAuditRecord, summary: SummaryMetrics) -> list[object]:
        return [
            self._pdf_name(row),
            row.original_url,
            self._category(row),
            None,
            None,
            self._axes_status(row),
            summary.run_timestamp.date(),
            self._tracker_bucket_value(row.grouped_results.pdf_ua_result),
            self._tracker_bucket_value(row.grouped_results.wcag_result),
            None,
            self._pdf_ua_notes(row),
            self._wcag_notes(row),
            "Not Started",
            self._adobe_status(row),
            self._remediation_notes(row),
            None,
            None,
            None,
            None,
        ]

    def _technical_row(self, row: DocumentAuditRecord) -> list[object]:
        return [
            row.original_url,
            row.final_url,
            row.retrieval_category.value,
            row.http_status,
            row.redirect_count,
            row.content_type,
            row.page_title,
            row.page_count,
            row.overall_result.value,
            row.grouped_results.hsbc_policy_result.value,
            row.failure_summary,
            row.failure_detail,
            row.remediation_guidance,
            row.manual_review_summary,
            row.notes,
        ]

    @staticmethod
    def _pdf_name(row: DocumentAuditRecord) -> str:
        if row.pdf_name:
            return row.pdf_name
        filename = Path(urlparse(row.original_url).path).stem
        return filename.replace("-", " ").replace("_", " ").title()

    @staticmethod
    def _category(row: DocumentAuditRecord) -> str | None:
        parts = [part for part in urlparse(row.original_url).path.split("/") if part]
        if "pdfs" in parts:
            index = parts.index("pdfs")
            if index + 1 < len(parts):
                return parts[index + 1].replace("-", " ").title()
        return None

    def _axes_status(self, row: DocumentAuditRecord) -> str:
        if row.retrieval_category != RetrievalCategory.DIRECT_FILE_OK:
            return "Blocked"
        if self._has_adobe_rule_failure(row, "Tagged PDF"):
            return "Blocked"
        return "Complete"

    @staticmethod
    def _tracker_bucket_value(status: CheckStatus) -> str | None:
        if status == CheckStatus.PASS:
            return "Pass"
        if status in {CheckStatus.FAIL, CheckStatus.NEEDS_MANUAL_REVIEW, CheckStatus.API_UNAVAILABLE}:
            return "Fail"
        return None

    def _pdf_ua_notes(self, row: DocumentAuditRecord) -> str | None:
        if row.retrieval_category != RetrievalCategory.DIRECT_FILE_OK:
            return "The source PDF could not be retrieved for accessibility testing."

        if self._has_adobe_rule_failure(row, "Tagged PDF"):
            return (
                "Basic requirement for accessible PDFs not met!\n\n"
                "The PDF does not contain an invisible structural layer in the form of tags and "
                "therefore does not meet the requirements for further accessibility testing."
            )

        if row.grouped_results.pdf_ua_result == CheckStatus.PASS:
            return "The file meets all machine-verifiable PDF/UA requirements."

        findings = self._bucket_findings(
            row,
            {
                "doc_title_present",
                "doc_language_present",
                "at_access_not_blocked",
                "figure_alt_present",
                "bookmarks_present_if_gt_3_pages",
                "lists_use_l",
                "tables_use_table_tag",
                "links_use_link_tag",
                "form_fields_have_tu",
                "footnotes_use_note",
                "adobe_full_check",
            },
            exclude_adobe_manual_only=True,
        )
        if findings:
            return " | ".join(findings[:4])
        return None

    def _wcag_notes(self, row: DocumentAuditRecord) -> str | None:
        if row.retrieval_category != RetrievalCategory.DIRECT_FILE_OK:
            return "The source PDF resolved to a not-found or non-PDF destination."

        if row.grouped_results.wcag_result == CheckStatus.PASS:
            return "The file meets all machine-verifiable WCAG requirements (Level A and AA)."

        if self._has_adobe_rule_status(row, "Color contrast", CheckStatus.NEEDS_MANUAL_REVIEW):
            return (
                "Adobe flagged color contrast for manual review. This typically indicates text/background "
                "contrast may not meet accessibility expectations."
            )

        findings = self._bucket_findings(
            row,
            {
                "url_resolves",
                "url_is_expected_file",
                "extractable_text_present",
                "colour_contrast_machine_check",
                "reading_order_machine_check",
                "adobe_full_check",
            },
            exclude_adobe_manual_only=False,
        )
        if findings:
            return " | ".join(findings[:4])
        return None

    def _adobe_status(self, row: DocumentAuditRecord) -> str:
        adobe_result = row.rule_results["adobe_full_check"].status
        if adobe_result in {CheckStatus.NA, CheckStatus.API_UNAVAILABLE}:
            return "Not Started"
        return "Complete"

    @staticmethod
    def _remediation_notes(row: DocumentAuditRecord) -> str | None:
        parts = [part for part in [row.remediation_guidance, row.manual_review_summary] if part]
        return " | ".join(parts) if parts else None

    @staticmethod
    def _clean_cell_value(value: object) -> object:
        if isinstance(value, str):
            return ILLEGAL_XLSX_CHAR_RE.sub("", value)
        return value

    @staticmethod
    def _adobe_findings(row: DocumentAuditRecord) -> list[dict]:
        raw = row.rule_results["adobe_full_check"].raw or {}
        detailed = raw.get("Detailed Report")
        if not isinstance(detailed, dict):
            return []
        findings: list[dict] = []
        for group in detailed.values():
            if isinstance(group, list):
                for finding in group:
                    if isinstance(finding, dict):
                        findings.append(finding)
        return findings

    def _has_adobe_rule_failure(self, row: DocumentAuditRecord, rule_name: str) -> bool:
        return self._has_adobe_rule_status(row, rule_name, CheckStatus.FAIL)

    def _has_adobe_rule_status(self, row: DocumentAuditRecord, rule_name: str, status: CheckStatus) -> bool:
        for finding in self._adobe_findings(row):
            rule = str(finding.get("Rule", "")).strip().lower()
            finding_status = str(finding.get("Status", "")).strip().lower()
            if rule == rule_name.lower():
                if status == CheckStatus.FAIL and finding_status == "failed":
                    return True
                if status == CheckStatus.NEEDS_MANUAL_REVIEW and finding_status == "needs manual check":
                    return True
        return False

    def _bucket_findings(
        self,
        row: DocumentAuditRecord,
        rule_ids: set[str],
        *,
        exclude_adobe_manual_only: bool,
    ) -> list[str]:
        findings: list[str] = []
        for rule_id in rule_ids:
            result = row.rule_results.get(rule_id)
            if not result:
                continue
            if result.status not in {CheckStatus.FAIL, CheckStatus.NEEDS_MANUAL_REVIEW}:
                continue
            if exclude_adobe_manual_only and rule_id in {"colour_contrast_machine_check", "reading_order_machine_check"}:
                continue
            title = RULE_DEFINITIONS[rule_id].title
            detail = result.evidence[0] if result.evidence else title
            findings.append(detail if detail != title else f"{title} needs attention.")
        return findings


def compute_summary(rows: list[DocumentAuditRecord]) -> SummaryMetrics:
    summary = SummaryMetrics(run_timestamp=datetime.utcnow())
    summary.total = len(rows)
    for row in rows:
        if row.retrieval_category != RetrievalCategory.DIRECT_FILE_OK:
            summary.unreachable_count += 1

        if row.overall_result == CheckStatus.PASS:
            summary.pass_count += 1
        elif row.overall_result == CheckStatus.FAIL:
            summary.fail_count += 1
        elif row.overall_result == CheckStatus.NEEDS_MANUAL_REVIEW:
            summary.manual_review_count += 1

        for rule_id, result in row.rule_results.items():
            if result.status == CheckStatus.FAIL:
                summary.per_rule_failures[rule_id] = summary.per_rule_failures.get(rule_id, 0) + 1

    return summary
