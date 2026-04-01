from io import BytesIO

from openpyxl import load_workbook

from app.models.domain import (
    CheckStatus,
    DocumentAuditRecord,
    GroupedResults,
    RetrievalCategory,
    RuleResult,
)
from app.services.adobe_cache import AdobeReportStore
from app.services.reporting import REPORT_COLUMNS, ReportBuilder, WorkbookTemplateResolver, compute_summary
from app.services.rule_catalog import RULE_DEFINITIONS


def _rule_result(rule_id: str, status: CheckStatus) -> RuleResult:
    definition = RULE_DEFINITIONS[rule_id]
    return RuleResult(
        rule_id=rule_id,
        theme=definition.theme,
        execution_mode=definition.execution_mode,
        status=status,
        evidence=["test evidence"] if status != CheckStatus.PASS else [],
        remediation=definition.remediation_template,
        confidence=1.0,
        source="test",
    )


def _record(status: CheckStatus) -> DocumentAuditRecord:
    results = {
        rule_id: _rule_result(rule_id, CheckStatus.PASS)
        for rule_id in RULE_DEFINITIONS
    }
    results["doc_title_present"] = _rule_result("doc_title_present", status)
    return DocumentAuditRecord(
        id=1,
        original_url="https://example.com/test.pdf",
        final_url="https://example.com/test.pdf",
        http_status="200",
        retrieval_category=RetrievalCategory.DIRECT_FILE_OK,
        overall_result=status,
        grouped_results=GroupedResults(
            pdf_ua_result=status,
            wcag_result=status,
            hsbc_policy_result=CheckStatus.PASS,
        ),
        rule_results=results,
        failure_summary="Document Title" if status != CheckStatus.PASS else "",
        failure_detail="Document Title: missing title metadata" if status != CheckStatus.PASS else "",
        remediation_guidance="Set document title metadata." if status != CheckStatus.PASS else "",
        manual_review_summary="",
        notes="",
    )


def test_compute_summary_counts_rule_failures() -> None:
    rows = [_record(CheckStatus.FAIL)]
    summary = compute_summary(rows)
    assert summary.total == 1
    assert summary.fail_count == 1
    assert summary.per_rule_failures["doc_title_present"] == 1


def test_report_builder_returns_workbook_bytes() -> None:
    rows = [_record(CheckStatus.PASS)]
    summary = compute_summary(rows)
    builder = ReportBuilder(WorkbookTemplateResolver(template_path=None))
    report = builder.build(rows, summary)
    assert report[:2] == b"PK"

    workbook = load_workbook(BytesIO(report))
    sheet = workbook["Sheet1"]
    assert workbook.sheetnames == ["Sheet1"]
    assert sheet.max_row == 3
    assert sheet["A1"].value == "Tracker Summary"
    assert sheet["A2"].value == "PDF Name"
    assert sheet["A3"].hyperlink.target == "https://example.com/test.pdf"
    assert sheet["B3"].value is None
    assert sheet["D3"].value == "Pass"
    assert sheet["E3"].value == "Pass"
    assert sheet.freeze_panes == "A3"
    assert any(str(range_ref).startswith("A1:") for range_ref in sheet.merged_cells.ranges)
    assert REPORT_COLUMNS[15] == "Original URL"


def test_report_builder_renders_review_and_pending_labels() -> None:
    manual_row = _record(CheckStatus.PASS)
    manual_row.grouped_results.wcag_result = CheckStatus.NEEDS_MANUAL_REVIEW
    manual_row.rule_results["colour_contrast_machine_check"] = _rule_result(
        "colour_contrast_machine_check", CheckStatus.NEEDS_MANUAL_REVIEW
    )
    manual_row.rule_results["reading_order_machine_check"] = _rule_result(
        "reading_order_machine_check", CheckStatus.NEEDS_MANUAL_REVIEW
    )
    manual_row.rule_results["adobe_full_check"] = _rule_result("adobe_full_check", CheckStatus.NEEDS_MANUAL_REVIEW)

    pending_row = _record(CheckStatus.PASS)
    pending_row.id = 2
    pending_row.original_url = "https://example.com/pending.pdf"
    pending_row.final_url = "https://example.com/pending.pdf"
    pending_row.grouped_results.wcag_result = CheckStatus.API_UNAVAILABLE
    pending_row.rule_results["colour_contrast_machine_check"] = _rule_result(
        "colour_contrast_machine_check", CheckStatus.API_UNAVAILABLE
    )
    pending_row.rule_results["reading_order_machine_check"] = _rule_result(
        "reading_order_machine_check", CheckStatus.API_UNAVAILABLE
    )
    pending_row.rule_results["adobe_full_check"] = _rule_result("adobe_full_check", CheckStatus.API_UNAVAILABLE)

    summary = compute_summary([manual_row, pending_row])
    builder = ReportBuilder(WorkbookTemplateResolver(template_path=None))
    report = builder.build([manual_row, pending_row], summary)

    workbook = load_workbook(BytesIO(report))
    sheet = workbook["Sheet1"]
    headers = [cell.value for cell in sheet[2]]
    header_index = {str(name): idx + 1 for idx, name in enumerate(headers) if name is not None}
    assert sheet.cell(3, header_index["WCAG Reults"]).value == "Needs Review"
    assert "human review" in sheet.cell(3, header_index["WCAG Notes"]).value
    assert sheet.cell(4, header_index["WCAG Reults"]).value == "Pending Adobe"
    assert "Re-run the Adobe step" in sheet.cell(4, header_index["WCAG Notes"]).value


def test_report_builder_maps_adobe_findings_into_wcag_remediation_notes() -> None:
    row = _record(CheckStatus.PASS)
    row.grouped_results.wcag_result = CheckStatus.NEEDS_MANUAL_REVIEW
    row.rule_results["adobe_full_check"] = _rule_result("adobe_full_check", CheckStatus.NEEDS_MANUAL_REVIEW)
    row.rule_results["adobe_full_check"].raw = {
        "Detailed Report": {
            "Document": [
                {"Rule": "Color contrast", "Status": "Needs manual check"},
                {"Rule": "Logical Reading Order", "Status": "Needs manual check"},
                {"Rule": "Title", "Status": "Failed"},
            ]
        }
    }

    summary = compute_summary([row])
    builder = ReportBuilder(WorkbookTemplateResolver(template_path=None))
    report = builder.build([row], summary)

    workbook = load_workbook(BytesIO(report))
    sheet = workbook["Sheet1"]
    headers = [cell.value for cell in sheet[2]]
    header_index = {str(name): idx + 1 for idx, name in enumerate(headers) if name is not None}
    remediation_notes = sheet.cell(3, header_index["Remediation Notes"]).value
    assert "WCAG 1.4.3 Contrast (Minimum)" in remediation_notes
    assert "WCAG 1.3.2 Meaningful Sequence" in remediation_notes
    assert "WCAG 2.4.2 Page Titled" in remediation_notes
    assert sheet.cell(3, header_index["WCAG Criteria Affected"]).value
    assert "Color contrast" in sheet.cell(3, header_index["WCAG Adobe Findings"]).value


def test_report_builder_loads_cached_adobe_findings_for_wcag_columns(tmp_path) -> None:
    row = _record(CheckStatus.PASS)
    row.rule_results["adobe_full_check"].raw = None
    store = AdobeReportStore(tmp_path)
    store.save(
        row.original_url,
        {
            "Detailed Report": {
                "Document": [
                    {"Rule": "Color contrast", "Status": "Needs manual check"},
                    {"Rule": "Title", "Status": "Failed"},
                ]
            }
        },
    )

    summary = compute_summary([row])
    builder = ReportBuilder(WorkbookTemplateResolver(template_path=None), adobe_report_store=store)
    report = builder.build([row], summary)

    workbook = load_workbook(BytesIO(report))
    sheet = workbook["Sheet1"]
    headers = [cell.value for cell in sheet[2]]
    header_index = {str(name): idx + 1 for idx, name in enumerate(headers) if name is not None}
    assert "WCAG 1.4.3 Contrast (Minimum)" in sheet.cell(3, header_index["WCAG Criteria Affected"]).value
    assert "Color contrast" in sheet.cell(3, header_index["WCAG Adobe Findings"]).value
